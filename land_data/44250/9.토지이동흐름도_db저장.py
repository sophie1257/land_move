# =====================================================================
#  이동정리현황 엑셀 → 토지이동연혁/소유자연혁 분리 · 저장 · DB적재
# =====================================================================

# [목적]
# - "이동정리현황" 엑셀 파일을 읽어 모든 값을 문자열로 변환(선행 0 보존)
# - 컬럼명에 '소유'가 포함된 컬럼은 소유자연혁(land_own), 나머지는 토지이동연혁(land_his)으로 분리
# - 분리된 두 DataFrame을 각각:
#   1) 텍스트 서식으로 엑셀 저장
#   2) MySQL DB에 적재 (스키마 자동 생성/갱신, 데이터 삽입 전 삭제)

# [입력 파일]
# - 우선순위:
#   ./1.data/out/이동정리현황_기간내.xlsx
#   ./1.data/in/*.xlsx 중 파일명에 "이동정리현황" 포함된 첫 파일 (fallback)

# [출력 파일]
# - ./1.data/out/토지이동연혁_split.xlsx
# - ./1.data/out/소유자연혁_split.xlsx

# [출력 DB]
# - DB: landmove (없으면 생성)
# - Table: land_his (토지이동연혁)
# - Table: land_own (소유자연혁)
# - id BIGINT AUTO_INCREMENT PRIMARY KEY 자동 생성
# - 컬럼명: 비영문/공백 등은 안전한 이름으로 치환
# - 타입 추론: 값 길이가 255자 초과면 TEXT, 아니면 VARCHAR(255)

# [실행 방법]
# > python <이파일이름>.py

# [의존성]
# - pandas
# - openpyxl
# - pymysql

# [주의]
# - DB_HOST/PORT/USER/PASS/NAME 은 코드 내 상수 또는 환경변수로 설정
# - 테이블이 존재할 경우, 없는 컬럼은 자동 추가 (VARCHAR(255))
# - 데이터 적재 시 기존 행은 모두 삭제 후 새 데이터 삽입

from pathlib import Path
import glob
import os
import re
from typing import Tuple, List

import pandas as pd
import pymysql
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers

# ============== 사용자 설정(필요시 수정) ==============
# 파일 검색 우선순위
CANDIDATE_PATHS = [
    Path("./1.data/out/이동정리현황_기간내.xlsx"),
    Path("./1.data/out/이동정리현황_기간내.xlsx"),
]

OUT_DIR = Path("./1.data/out")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 출력 파일명
OUT_XLSX_LAND_HIS  = OUT_DIR / "토지이동연혁_split.xlsx"
OUT_XLSX_LAND_OWN  = OUT_DIR / "소유자연혁_split.xlsx"

# DB 접속정보 (환경변수가 있으면 우선 사용)
DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_PORT = int(os.getenv("DB_PORT", "3306"))
DB_USER = os.getenv("DB_USER", "root")
DB_PASS = os.getenv("DB_PASS", "1234")
DB_NAME = os.getenv("DB_NAME", "landmove")

TABLE_HIS = "land_his"
TABLE_OWN = "land_own"

# ============== 유틸 ==============
def find_input_file() -> Path:
    """우선순위 경로 → 없으면 ./1.data/in/*.xlsx 중 '이동정리현황' 포함 파일 첫번째"""
    for p in CANDIDATE_PATHS:
        if p.is_file():
            return p

    # fallback: 44250 폴더 전체 스캔
    for p in glob.glob("./1.data/in/*.xlsx"):
        if "이동정리현황" in Path(p).name:
            return Path(p)

    raise FileNotFoundError("입력 엑셀을 찾지 못했습니다. ./out 또는 ./1.data/in/ 경로를 확인하세요.")

def read_excel_as_text(path: Path) -> pd.DataFrame:
    """엑셀을 모든 값을 문자열로 읽기 (선행 0 보존). 빈값은 빈문자열로 통일."""
    # dtype=str로 읽더라도 NaN이 생길 수 있어 후처리
    df = pd.read_excel(path, dtype=str)
    df = df.fillna("")  # 결측을 빈문자열로
    # 열 이름 좌우 공백 제거
    df.columns = [str(c).strip() for c in df.columns]
    return df

def split_by_owner_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    컬럼명에 '소유'가 들어가면 소유자연혁(land_own)으로, 나머지는 토지이동연혁(land_his)으로 분리
    행(row) 수는 동일, 컬럼만 분리
    """
    owner_mask = df.columns.to_series().str.contains("소유", na=False)
    own_cols = df.columns[owner_mask].tolist()
    his_cols = df.columns[~owner_mask].tolist()

    df_own = df[own_cols].copy() if own_cols else pd.DataFrame(index=df.index)
    df_his = df[his_cols].copy() if his_cols else pd.DataFrame(index=df.index)

    # 인덱스 초기화(필요시)
    df_own.reset_index(drop=True, inplace=True)
    df_his.reset_index(drop=True, inplace=True)
    return df_his, df_own

def save_excel_text(df: pd.DataFrame, out_path: Path):
    """openpyxl로 모든 셀을 텍스트 서식(number_format='@')으로 저장."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더
    ws.append(list(df.columns))

    # 데이터
    for _, row in df.iterrows():
        ws.append([("" if pd.isna(v) else str(v)) for v in row.tolist()])

    # 모든 셀 텍스트 서식 적용
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.number_format = numbers.FORMAT_TEXT  # "@"

    wb.save(out_path)

def infer_mysql_type(series: pd.Series) -> str:
    """
    간단 타입 추론:
      - 열 전체에서 255자 초과 값이 있는 경우 TEXT
      - 그 외 VARCHAR(255)
    모든 값은 문자열로 들어온다고 가정.
    """
    max_len = series.astype(str).map(len).max() if len(series) else 0
    if pd.isna(max_len):
        max_len = 0
    return "TEXT" if max_len and max_len > 255 else "VARCHAR(255)"

def ensure_database_and_table(conn, table: str, df: pd.DataFrame):
    """DB/테이블 생성 보장. PRIMARY KEY는 자동 증가 id 추가."""
    with conn.cursor() as cur:
        # DB 생성
        cur.execute(f"CREATE DATABASE IF NOT EXISTS `{DB_NAME}` CHARACTER SET utf8mb4 COLLATE utf8mb4_general_ci;")
        cur.execute(f"USE `{DB_NAME}`;")

        # 테이블 존재 체크
        cur.execute("SELECT COUNT(*) FROM information_schema.tables WHERE table_schema=%s AND table_name=%s;", (DB_NAME, table))
        exists = cur.fetchone()[0] > 0

        if not exists:
            # 스키마 생성
            cols_sql: List[str] = ["`id` BIGINT NOT NULL AUTO_INCREMENT"]
            for col in df.columns:
                col_name = re.sub(r"[^\w가-힣_]", "_", str(col))  # 안전한 컬럼명으로
                col_type = infer_mysql_type(df[col])
                cols_sql.append(f"`{col_name}` {col_type} NULL")
            cols_sql.append("PRIMARY KEY (`id`)")
            create_sql = f"CREATE TABLE `{table}` (\n  " + ",\n  ".join(cols_sql) + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;"
            cur.execute(create_sql)
        else:
            # 존재한다면 컬럼 추가 필요 여부 체크(간단하게 없는 컬럼만 VARCHAR(255)로 추가)
            cur.execute(f"SHOW COLUMNS FROM `{table}`;")
            existing_cols = {row[0] for row in cur.fetchall()}
            for col in df.columns:
                safe_col = re.sub(r"[^\w가-힣_]", "_", str(col))
                if safe_col not in existing_cols:
                    cur.execute(f"ALTER TABLE `{table}` ADD COLUMN `{safe_col}` {infer_mysql_type(df[col])} NULL;")

    conn.commit()

def clear_table(conn, table: str):
    """테이블의 기존 데이터를 모두 삭제"""
    with conn.cursor() as cur:
        cur.execute(f"DELETE FROM `{table}`;")   # 또는 TRUNCATE TABLE `{table}`;
    conn.commit()
    print(f"[RESET] Cleared all rows in {table}")


def insert_dataframe(conn, table: str, df: pd.DataFrame):
    """DataFrame을 INSERT. 기존 데이터는 먼저 삭제."""
    if df.empty:
        print(f"[INFO] {table}: 비어 있어 적재 생략")
        return

    # 기존 데이터 삭제
    clear_table(conn, table)

    # 컬럼명 정규화
    safe_cols = [re.sub(r"[^\w가-힣_]", "_", str(c)) for c in df.columns]
    df2 = df.copy()
    df2.columns = safe_cols

    # 문자열 변환
    df2 = df2.fillna("").astype(str)

    cols_clause = ", ".join([f"`{c}`" for c in df2.columns])
    placeholders = ", ".join(["%s"] * len(df2.columns))
    sql = f"INSERT INTO `{table}` ({cols_clause}) VALUES ({placeholders})"

    with conn.cursor() as cur:
        cur.executemany(sql, df2.values.tolist())
    conn.commit()
    print(f"[OK] Inserted {len(df2)} rows into {table}")


def main():
    # 1) 입력 로딩
    in_path = find_input_file()
    print(f"[INFO] 입력 파일: {in_path}")
    df = read_excel_as_text(in_path)
    print(f"[INFO] 원본 shape: {df.shape}")

    # 2) 분리
    df_his, df_own = split_by_owner_columns(df)
    print(f"[INFO] 토지이동연혁(land_his) shape: {df_his.shape}")
    print(f"[INFO] 소유자연혁(land_own) shape: {df_own.shape}")

    # 3) 엑셀 저장(모든 셀 텍스트)
    save_excel_text(df_his, OUT_XLSX_LAND_HIS)
    save_excel_text(df_own, OUT_XLSX_LAND_OWN)
    print(f"[OK] 저장: {OUT_XLSX_LAND_HIS}")
    print(f"[OK] 저장: {OUT_XLSX_LAND_OWN}")

    # 4) DB 적재
    conn = pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASS,
        autocommit=False,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.Cursor,
    )

    try:
        ensure_database_and_table(conn, TABLE_HIS, df_his)
        ensure_database_and_table(conn, TABLE_OWN, df_own)
        insert_dataframe(conn, TABLE_HIS, df_his)
        insert_dataframe(conn, TABLE_OWN, df_own)
    finally:
        conn.close()
    print("[DONE] 엑셀 분리 + DB 적재 완료")

if __name__ == "__main__":
    main()
