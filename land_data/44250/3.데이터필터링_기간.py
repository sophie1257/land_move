# =========================================================
#  기간 필터링 · 집계 · 엑셀저장 (YYYYMMDD: 20250101~20250630)
# =========================================================

# [목적]
# - 3개 원본(엑셀 2개 + CSV 1개)을 읽어 공통 날짜 컬럼으로
#   기간내/기간외 분리하고, 기간외 존재 여부 출력
# - 토지(임야)기본/이동정리현황: 기간내 데이터에 대해 이동사유(우선) 또는
#   이동종목 기준 집계표를 콘솔에 출력
# - 각 파일의 기간내/기간외 데이터를 모두 텍스트 서식(@)으로 엑셀 저장

# [입력 파일]  (스크립트와 동일 루트 기준)
# - ./1.data/out/토지(임야)기본_필지코드추가.xlsx
# - .
# /1.data/out/토지이동정리현황_필지코드추가.xlsx
# - ./44250/1.data/in/일반용조서(말소용).csv

# [출력 파일]  (없으면 생성, 기존 파일 있으면 타임스탬프 부여 저장)
# - 44250/1.data/out/토지(임야)기본_기간내.xlsx
# - 44250/1.data/out/토지(임야)기본_기간외.xlsx
# - 44250/1.data/out/이동정리현황_기간내.xlsx
# - 44250/1.data/out/이동정리현황_기간외.xlsx
# - 44250/1.data/out/일반용조서(말소용)_기간내.xlsx
# - 44250/1.data/out/일반용조서(말소용)_기간외.xlsx

# [실행 방법]
# > python 3.데이터필터링_기간.py
# -모듈설치: pandas, openpyxl

from pathlib import Path
from datetime import datetime
import re
import pandas as pd
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# -----------------------------
# 공통 설정
# -----------------------------
BASE = Path("44250").resolve()
DATE_RANGE = ("20250101", "20250630")  # [시작, 종료] (YYYYMMDD)

# 파일 경로
P_LAND_BASIC  = BASE /"1.data"/ "out" / "토지(임야)기본_필지코드추가.xlsx"
P_MOVE_STATUS = BASE /"1.data"/ "out" / "토지이동정리현황_필지코드추가.xlsx"
P_MALSO_CSV   = BASE /"1.data"/"in"/ "일반용조서(말소용).csv"

# -----------------------------
# 공통 유틸
# -----------------------------
def _strip_digits(s: str) -> str:
    """숫자만 남기기"""
    return re.sub(r"\D+", "", str(s)) if pd.notna(s) else ""

def _normalize_yyyymmdd(s: str) -> str:
    """
    날짜 문자열에서 숫자만 추출 → 8자리(YYYYMMDD)만 사용.
    잘리거나 부족하면 빈문자.
    """
    ds = _strip_digits(s)
    return ds[:8] if len(ds) >= 8 else ""

def _find_first_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """후보 리스트에서 처음으로 존재하는 컬럼명을 반환(대소문자 무시)"""
    lower_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None

def _read_excel_all_text(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, dtype=str, engine="openpyxl")

def _read_csv_guess_encoding(path: Path) -> pd.DataFrame:
    for enc in ["utf-8", "utf-8-sig", "cp949"]:
        try:
            return pd.read_csv(path, dtype=str, encoding=enc)
        except Exception:
            continue
    return pd.read_csv(path, dtype=str)

def _filter_by_date(df: pd.DataFrame, date_col: str):
    """
    date_col 기준으로 기간 내/외 분리
    return: (기간내_df, 기간외_df, 기간외_존재여부)
    """
    tmp = df.copy()
    tmp["_YMD"] = tmp[date_col].map(_normalize_yyyymmdd)

    in_mask = (tmp["_YMD"] >= DATE_RANGE[0]) & (tmp["_YMD"] <= DATE_RANGE[1])
    in_df = tmp.loc[in_mask].drop(columns=["_YMD"])
    out_df = tmp.loc[~in_mask].drop(columns=["_YMD"])
    has_out = len(out_df) > 0
    return in_df, out_df, has_out

def _groupby_count(df: pd.DataFrame, by_col: str, label: str) -> pd.DataFrame:
    g = df.groupby(by_col, dropna=False).size().reset_index(name="건수")
    g = g.rename(columns={by_col: label})
    g[label] = g[label].fillna("(미기재)")
    return g.sort_values("건수", ascending=False)

def _print_section(title: str):
    print("\n" + "="*70)
    print(f"[{title}]")
    print("="*70)

def _save_excel_all_text(df: pd.DataFrame, out_path: Path):
    """
    DataFrame을 모든 셀 텍스트 서식(@)으로 엑셀 저장
    (엑셀이 열 때 선행 0 보존)
    """
    target = out_path
    if target.exists():  # 파일 잠금/중복 대비 타임스탬프 부여
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = out_path.with_name(out_path.stem + f"_{ts}" + out_path.suffix)

    wb = Workbook()
    ws = wb.active
    ws.title = "기간_데이터"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 모든 셀 텍스트 서식 적용
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.number_format = numbers.FORMAT_TEXT  # "@"
            if cell.value is None:
                cell.value = ""  # 빈칸 통일

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    print(f"[저장] {target}")

# -----------------------------
# 개별 파일 처리
# -----------------------------
def process_land_basic():
    _print_section("토지(임야)기본 — 기간외 여부/집계/엑셀 저장")
    df = _read_excel_all_text(P_LAND_BASIC)

    # 날짜/사유 후보 (파일별 명칭 편차 흡수)
    date_candidates = ["토지이동일자", "정리일자", "cre_ymd"]
    reason_candidates = ["토지이동사유", "토지이동종목", "land_mov_rsn"]

    date_col = _find_first_col(df, date_candidates)
    if not date_col:
        raise ValueError(f"[오류] 날짜열을 찾을 수 없습니다. 후보={date_candidates}\n현재 열: {list(df.columns)}")

    in_df, out_df, has_out = _filter_by_date(df, date_col)
    print(f"- 기간 외 자료 존재 여부: {'예' if has_out else '아니오'} (전체 {len(df):,}건 / 기간내 {len(in_df):,}건 / 기간외 {len(out_df):,}건)")

    # 사유 우선, 없으면 종목 대체
    reason_col = _find_first_col(in_df, ["토지이동사유", "land_mov_rsn"]) \
                 or _find_first_col(in_df, ["토지이동종목", "land_mov_type"])
    if reason_col:
        agg = _groupby_count(in_df, reason_col, "이동사유/종목")
        print("\n[이동사유/종목별 집계]")
        print(agg.to_string(index=False))
    else:
        print("\n[경고] 이동사유/종목 컬럼을 찾지 못했습니다.")

    # 저장: 기간내 + 기간외(추가)
    out_in  = BASE / "1.data/out" / "토지(임야)기본_기간내.xlsx"
    out_out = BASE / "1.data/out" / "토지(임야)기본_기간외.xlsx"
    _save_excel_all_text(in_df, out_in)
    if has_out:
        _save_excel_all_text(out_df, out_out)

def process_move_status():
    _print_section("이동정리현황 — 기간외 여부/집계/엑셀 저장")
    df = _read_excel_all_text(P_MOVE_STATUS)

    date_candidates = ["정리일자", "토지이동일자", "cre_ymd"]
    # 사유 우선, 없으면 종목
    reason_candidates = ["토지이동사유", "land_mov_rsn", "토지이동종목", "land_mov_type"]

    date_col = _find_first_col(df, date_candidates)
    if not date_col:
        raise ValueError(f"[오류] 날짜열을 찾을 수 없습니다. 후보={date_candidates}\n현재 열: {list(df.columns)}")

    in_df, out_df, has_out = _filter_by_date(df, date_col)
    print(f"- 기간 외 자료 존재 여부: {'예' if has_out else '아니오'} (전체 {len(df):,}건 / 기간내 {len(in_df):,}건 / 기간외 {len(out_df):,}건)")

    reason_col = _find_first_col(in_df, ["토지이동사유", "land_mov_rsn"]) \
                 or _find_first_col(in_df, ["토지이동종목", "land_mov_type"])
    if reason_col:
        agg = _groupby_count(in_df, reason_col, "이동사유/종목")
        print("\n[이동사유/종목별 집계]")
        print(agg.to_string(index=False))
    else:
        print("\n[경고] 이동사유/종목 컬럼을 찾지 못했습니다.")

    # 저장: 기간내 + 기간외
    out_in  = BASE /"1.data"/ "out" / "이동정리현황_기간내.xlsx"
    out_out = BASE /"1.data"/ "out" / "이동정리현황_기간외.xlsx"
    _save_excel_all_text(in_df, out_in)
    if has_out:
        _save_excel_all_text(out_df, out_out)

def process_malso_csv():
    """
    일반용조서(말소용).csv
    - 기간 외 여부, 기간내/기간외 데이터 저장(집계 없음)
    """
    _print_section("일반용조서(말소용) — 기간외 여부/엑셀 저장")
    df = _read_csv_guess_encoding(P_MALSO_CSV)

    date_candidates = ["토지이동일자", "정리일자", "cre_ymd"]
    date_col = _find_first_col(df, date_candidates)
    if not date_col:
        raise ValueError(f"[오류] 날짜열을 찾을 수 없습니다. 후보={date_candidates}\n현재 열: {list(df.columns)}")

    in_df, out_df, has_out = _filter_by_date(df, date_col)
    print(f"- 기간 외 자료 존재 여부: {'예' if has_out else '아니오'} (전체 {len(df):,}건 / 기간내 {len(in_df):,}건 / 기간외 {len(out_df):,}건)")

    # 저장: 기간내 + 기간외(추가)
    out_in  = BASE /"1.data"/ "out" / "일반용조서(말소용)_기간내.xlsx"
    out_out = BASE /"1.data"/ "out" / "일반용조서(말소용)_기간외.xlsx"
    _save_excel_all_text(in_df, out_in)
    if has_out:
        _save_excel_all_text(out_df, out_out)

# -----------------------------
# 실행
# -----------------------------
def main():
    print("[INFO] 입력 파일")
    print(" - 토지(임야)기본:", P_LAND_BASIC)
    print(" - 이동정리현황  :", P_MOVE_STATUS)
    print(" - 일반용조서(CSV):", P_MALSO_CSV)

    process_land_basic()
    process_move_status()
    process_malso_csv()

if __name__ == "__main__":
    main()
