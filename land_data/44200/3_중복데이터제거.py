"""
토지이동연혁 / 소유자변경이력 엑셀 → 중복 제거 후 별도 저장
- 파일명에서 행정구역(5자리)과 기간(YYMMDD-YYMMDD) 자동 추출 및 보정
- 전체 셀 TEXT 서식으로 저장 (선행 0 보존)

입력(예시):
  ./1.data/out/44200_20240102-20250630_토지이동연혁.xlsx
  ./1.data/out/44200_24010102-250630_소유자변경이력.xlsx

출력(요청 형식):
  ./1.data/out/44200_240101-250630_토지이동연혁_중복제거.xlsx
  ./1.data/out/44200_240101-250630_소유자변경이력_중복제거.xlsx

실행:
  python 3. 중복데이터제거.py
"""

from pathlib import Path
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers

# -----------------------------
# 사용자 입력 경로
# -----------------------------
BASE = Path("./1.data/out").resolve()

SRC_CORE  = BASE / "44200_20240102-20250630_토지이동연혁.xlsx"
SRC_OWNER = BASE / "44200_24010102-20250630_소유자변경이력.xlsx"

OUT_DIR = BASE
OUT_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------------
# 파일명에서 code5/기간 추출 & 보정
# -----------------------------
def extract_code5_and_period(p: Path) -> tuple[str, str]:
    """
    파일명에서 행정구역 5자리와 기간을 추출.
    """
    name = p.stem  # 확장자 제거
    # 행정구역 5자리
    m_code = re.search(r"\b(\d{5})\b", name)
    code5 = m_code.group(1) if m_code else "00000"

    # 기간 후보: 하이픈으로 구분된 두 덩어리
    # 왼쪽 6~8자리, 오른쪽 6~8자리 숫자를 폭넓게 허용
    m_period = re.search(r"(\d{6,8})-(\d{6,8})", name)
    if not m_period:
        # fallback: 언더스코어 구분 등 예외 처리
        m_period = re.search(r"(\d{6,8})[_](\d{6,8})", name)
    if m_period:
        left, right = m_period.group(1), m_period.group(2)
    else:
        # 못 찾으면 기본값
        left, right = "000000", "000000"

    def to_yymmdd(s: str) -> str:
        # 8자리(YYYYMMDD) → YYMMDD, 6자리(YYMMDD)는 그대로, 그 외는 6자리로 padding
        s = re.sub(r"\D", "", s)
        if len(s) == 8:
            return s[2:]     # YYYYMMDD -> YYMMDD
        elif len(s) == 6:
            return s
        elif len(s) > 6:
            return s[-6:]
        else:
            return s.zfill(6)

    period = f"{to_yymmdd(left)}-{to_yymmdd(right)}"
    return code5, period

# -----------------------------
# 저장 유틸(모든 셀 TEXT)
# -----------------------------
def save_excel_text(df: pd.DataFrame, path: Path, sheetname="Sheet1"):
    # 문자열화 + 좌우 공백 제거(중복 판단 정교화)
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()

    wb = Workbook()
    ws = wb.active
    ws.title = sheetname

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 전체 셀 텍스트 서식
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                          min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.number_format = numbers.FORMAT_TEXT

    # 간단 열 너비 조정
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        width = max(10, min(80, max(len(str(c.value)) if c.value else 0 for c in col_cells) + 2))
        ws.column_dimensions[letter].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[OK] 저장 완료: {path}")

# -----------------------------
# 1) 토지이동연혁: 중복 제거 저장
# -----------------------------
df_core = pd.read_excel(SRC_CORE, dtype=str)
for c in df_core.columns:
    df_core[c] = df_core[c].astype(str).str.strip()
df_core = df_core.drop_duplicates(keep="first").reset_index(drop=True)

code5_core, period_core = extract_code5_and_period(SRC_CORE)
out_core = OUT_DIR / f"{code5_core}_{period_core}_토지이동연혁_중복제거.xlsx"
save_excel_text(df_core, out_core, sheetname="토지이동연혁")

# -----------------------------
# 2) 소유자변경이력: 중복 제거 저장
# -----------------------------
df_owner = pd.read_excel(SRC_OWNER, dtype=str)
for c in df_owner.columns:
    df_owner[c] = df_owner[c].astype(str).str.strip()
df_owner = df_owner.drop_duplicates(keep="first").reset_index(drop=True)

code5_owner, period_owner = extract_code5_and_period(SRC_OWNER)
out_owner = OUT_DIR / f"{code5_owner}_{period_owner}_소유자변경이력_중복제거.xlsx"
save_excel_text(df_owner, out_owner, sheetname="소유자변경이력")
