"""
==========================================
토지이동정리현황 CSV 정제
==========================================

[목적]
- 세 개의 CSV 파일을 로드하여 병합
- 필지코드(19자리) 생성
- 지목, 소유구분, 이동종목 등의 값 정제
- 지정한 기간(20240102~20250630) 기준으로 기간내/기간외 자료 분리
- 모든 셀을 텍스트 서식으로 지정한 Excel 파일로 저장

[입력 파일]-경로지정
./1.data/in/토지이동정리현황(소유권포함)(2024_01).csv
./1.data/in/토지이동정리현황(소유권포함)(2024_07).csv
./1.data/in/토지이동정리현황(소유권포함)(2025_01).csv

[출력 파일]-경로지정
./1.data/out/20240102_20250630_이동정리현황_기간내.xlsx
./1.data/out/20240102_20250630_이동정리현황_기간외.xlsx

[실행 방법]
터미널에서 실행:
    python 1_pnu코드정제.py

"""

from pathlib import Path
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers

# -----------------------------
# 경로/입출력 설정
# -----------------------------
BASE = Path("./1.data/in").resolve()
IN_FILES = [
    BASE / "토지이동정리현황(소유권포함)(2024_01).csv",
    BASE / "토지이동정리현황(소유권포함)(2024_07).csv",
    BASE / "토지이동정리현황(소유권포함)(2025_01).csv",
]
OUT_DIR = BASE / "../out"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 시군구코드
DISTRICT_CODE = "44200"
#기간
DATE_START = "20240102"
DATE_END   = "20250630"
PERIOD_TAG = f"{DATE_START}_{DATE_END}"

OUT_IN   = OUT_DIR / f"{DISTRICT_CODE}_기간내_자료.xlsx"
OUT_OUT  = OUT_DIR / f"{DISTRICT_CODE}_기간외_자료.xlsx"

# 필수 컬럼
REQUIRED_COLS = [
    "지역코드", "대장구분",
    "이동전_지번", "이동후_지번",
    "이동전_지목", "이동후_지목",
    "현재_소유구분",
    "토지이동종목",
    "정리일자",
]

# 삭제 대상
DROP_COLS = [
    "일련번호", "지역코드", "대장구분", "이동전_지번", "이동후_지번",
    "신청_소유구분", "신청_소유자명", "신청_소유자주소",
    "공시지가", "공시지가_수시", "전년지가", "전년지가_수시",
    "2년전지가", "2년전지가_수시", "3년전지가", "3년전지가_수시",
    "4년전지가", "4년전지가_수시",
]

MOVE_MAP = {
    "분할(임야대장)": "20",
    "분할(토지대장)": "20",
    "합병(토지대장)": "30",
    "지목변경(토지대장)": "40",
    "등록사항정정(토지대장)": "10",
}

# -----------------------------
# 유틸
# -----------------------------
def digits_only(s: str) -> str:
    return re.sub(r"\D", "", "" if s is None else str(s)) #숫자만추출(이동전_지목, 이동후_지목, 현재_소유구분)

def norm_date8(s: str) -> str:  #날짜 8자리
    ss = digits_only(s)
    return ss if len(ss) == 8 else ""

def make_pnu(region: str, ledger: str, jibun: str) -> str: #PNU 19자리 생성
    reg = digits_only(region).zfill(10)
    led = ("" if ledger is None else str(ledger))[:1]
    jbn = digits_only(jibun).zfill(8)
    return reg + led + jbn

def to2digits(s: str) -> str: #현재_소유구분(숫자코드 2자리)
    )    d = digits_only(s)
    return d.zfill(2)[-2:] if d else ""

def save_excel(df: pd.DataFrame, path: Path, sheetname="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.number_format = numbers.FORMAT_TEXT
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        width = max(10, min(80, max(len(str(c.value)) if c.value else 0 for c in col_cells) + 2))
        ws.column_dimensions[col_letter].width = width
    wb.save(path)
    print(f"[OK] 저장 완료: {path}")

# -----------------------------
# 1) CSV 로딩/병합
# -----------------------------
def read_csv_keep_strings(path: Path) -> pd.DataFrame:
    for enc in ("utf-8"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc)
        except Exception:
            continue
    raise RuntimeError(f"CSV 인코딩 실패: {path}")

dfs = [read_csv_keep_strings(p) for p in IN_FILES]
df = pd.concat(dfs, ignore_index=True)

# -----------------------------
# 2) 컬럼 검증
# -----------------------------
missing = [c for c in REQUIRED_COLS if c not in df.columns]
if missing:
    raise ValueError(f"필수 컬럼이 없습니다: {missing}")

# -----------------------------
# 3) 정제
# -----------------------------
df["이동전_필지코드"] = df.apply(lambda r: make_pnu(r["지역코드"], r["대장구분"], r["이동전_지번"]), axis=1)
df["이동후_필지코드"] = df.apply(lambda r: make_pnu(r["지역코드"], r["대장구분"], r["이동후_지번"]), axis=1)
df["이동전_지목"] = df["이동전_지목"].map(to2digits)
df["이동후_지목"] = df["이동후_지목"].map(to2digits)
df["현재_소유구분"] = df["현재_소유구분"].map(lambda s: digits_only(s).lstrip("0") if s else "")
df["토지이동종목"] = df["토지이동종목"].map(lambda s: MOVE_MAP.get(str(s).strip(), digits_only(str(s)) or "")).astype(str)

df["_DATE8_"] = df["정리일자"].map(norm_date8)
df_in  = df[(df["_DATE8_"] >= DATE_START) & (df["_DATE8_"] <= DATE_END)].copy()
df_out = df.drop(df_in.index).copy()
df_in.drop(columns=["_DATE8_"], inplace=True)
df_out.drop(columns=["_DATE8_"], inplace=True)

drop_now = [c for c in DROP_COLS if c in df.columns]
df_in.drop(columns=drop_now, inplace=True, errors="ignore")
df_out.drop(columns=drop_now, inplace=True, errors="ignore")

front = ["이동전_필지코드", "이동후_필지코드"]
df_in  = df_in.reindex(columns=front + [c for c in df_in.columns if c not in front])
df_out = df_out.reindex(columns=front + [c for c in df_out.columns if c not in front])

for c in df_in.columns:  df_in[c]  = df_in[c].astype(str)
for c in df_out.columns: df_out[c] = df_out[c].astype(str)

# -----------------------------
# 4) 저장
# -----------------------------
save_excel(df_in, OUT_IN,  "44200_240101-250631_기간내_자료.xlsx")
save_excel(df_out, OUT_OUT, "44200_240101-250631_기간외_자료.xlsx")
